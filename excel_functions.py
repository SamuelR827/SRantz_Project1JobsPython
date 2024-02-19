""" This module handles the processing of the Excel worksheet data. """
import sqlite3

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from database_functions import insert_worksheet_data_to_database


def load_job_workbook() -> Worksheet:
    """ This function loads the active job Excel worksheet from the project
    and returns it."""
    job_workbook = load_workbook('Sprint3Data.xlsx')
    job_worksheet = job_workbook.active
    print(type(job_worksheet))
    return job_worksheet


def add_excel_job_data(cursor: sqlite3.Cursor, job_worksheet: Worksheet) -> None:
    """ This functions goes through each row in the specified worksheet passed as a parameter. For each
    row find the desired data in each column. Creates a tuple out of the desired data and calls the database
    function to insert the Excel data into the database."""
    # variable to get the amount of rows in the Excel sheet
    row_count = job_worksheet.max_row
    # loop through each row in the sheet starting at two for the first job row
    for row in range(2, row_count):
        company_name = job_worksheet.cell(row=row, column=1).value
        posted_ago = job_worksheet.cell(row=row, column=2).value
        job_id = job_worksheet.cell(row=row, column=3).value
        location = job_worksheet.cell(row=row, column=5).value
        salary_min = job_worksheet.cell(row=row, column=8).value
        salary_max = job_worksheet.cell(row=row, column=7).value
        salary_rate = job_worksheet.cell(row=row, column=9).value
        job_name = job_worksheet.cell(row=row, column=10).value
        # empty values for the database since Excel sheet doesn't have this data
        job_remote = 'N/A'
        job_description = 'N/A'
        job_worksheet_data = (job_id, job_name, company_name, job_description, location,
                              job_remote, posted_ago, salary_min, salary_max, salary_rate)
        insert_worksheet_data_to_database(cursor, job_worksheet_data)
